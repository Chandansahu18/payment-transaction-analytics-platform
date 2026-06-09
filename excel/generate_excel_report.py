import logging
import sys
from dataclasses import dataclass
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ingestion.db_utils import get_db_connection

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

OUTPUT_FILE = Path(__file__).parent / "Payment_Transaction_Analytics_Dashboard.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center")

def _fetch_df(query: str, conn) -> pd.DataFrame:
    with conn.cursor() as cur:
        cur.execute(query)
        columns = [desc[0] for desc in cur.description]
        return pd.DataFrame(cur.fetchall(), columns=columns)

def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

@dataclass
class SheetConfig:
    label: str
    query: str
    max_rows: int | None = None

    def qualified_query(self) -> str:
        q = self.query.rstrip(";")
        return f"{q} LIMIT {self.max_rows}" if self.max_rows else q

SHEETS = [
    SheetConfig("Cohort_Data", "SELECT * FROM public.cohort_analysis"),
    SheetConfig("Segments_Data", "SELECT * FROM public.fraud_customer_segments"),
    SheetConfig("Hourly_Data", "SELECT * FROM public.hourly_fraud_trends"),
    SheetConfig("Merchant_Data", "SELECT * FROM public.merchant_risk_profiling"),
    SheetConfig("Velocity_Data", "SELECT * FROM public.velocity_anomaly_detection"),
    SheetConfig("Raw_Transactions", "SELECT * FROM raw.transactions"),
]

def _write_sheet(wb: Workbook, df: pd.DataFrame, title: str):
    ws = wb.create_sheet(title=title)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = HEADER_FONT if r_idx == 1 else BODY_FONT
            cell.fill = HEADER_FILL if r_idx == 1 else PatternFill()
            cell.border = THIN_BORDER
            cell.alignment = CENTER
    n_cols = len(df.columns)
    ws.auto_filter.ref = f"A1:{_col_letter(n_cols)}{len(df) + 1}"
    for i, col in enumerate(df.columns, 1):
        data_len = df[col].astype(str).map(len).max() or 0
        ws.column_dimensions[_col_letter(i)].width = min(max(data_len, len(col)) + 2, 30)

def main():
    logger.info("Starting Excel data export")

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SET search_path TO staging, intermediate, marts, public")

        wb = Workbook()
        wb.remove(wb.worksheets[0])

        for cfg in SHEETS:
            try:
                df = _fetch_df(cfg.qualified_query(), conn)
                for col in df.select_dtypes(include=["datetimetz"]):
                    df[col] = df[col].dt.tz_localize(None)
                if "view_generated_at" in df.columns:
                    df = df.drop(columns=["view_generated_at"])
                if "cohort_month" in df.columns:
                    df["cohort_month"] = pd.to_datetime(df["cohort_month"]).dt.strftime("%Y-%m-%d")
                if "transaction_ts" in df.columns:
                    df["transaction_ts"] = pd.to_datetime(df["transaction_ts"]).dt.strftime("%Y-%m-%d %H:%M")
                _write_sheet(wb, df, cfg.label)
                logger.info("%-20s (%s rows)", cfg.label, f"{len(df):,}")
            except Exception:
                logger.exception("%s failed, skipping", cfg.label)

        wb.save(str(OUTPUT_FILE))

    logger.info("Done! data saved to %s", OUTPUT_FILE)

if __name__ == "__main__":
    main()